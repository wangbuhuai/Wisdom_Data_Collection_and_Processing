# Created by Dayu Wang (dwang@stchas.edu) on 2022-03-09

# Last updated by Dayu Wang (dwang@stchas.edu) on 2022-03-09


from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from os import listdir
from pandas import read_excel
from re import sub
from tkinter import Tk, filedialog


def extract(tokens):
    """ Extracts first name and last name from the tokens of a split full name.
        :param tokens: tokens split from a full name
        :type tokens: List[str]
        :return: a dictionary containing extracted first name, last name, and whether the result is reliable
        :rtype: Dict
    """
    result = {}

    # Create a new token list consists of all the name tokens that are greater than one character.
    effective_tokens = []
    for token in tokens:
        if len(token) > 1:
            effective_tokens.append(token)

    # Assigns the first two tokens in effective tokens as first name and last name.
    result["first_name"] = effective_tokens[0]
    result["last_name"] = effective_tokens[1]

    # If there are only two effective tokens, then the result is reliable.  Otherwise, the result is unreliable.
    result["reliable"] = len(effective_tokens) == 2

    return result


def main():
    # Open the database containing the full search names of CEOs.
    Tk().withdraw()
    ceo_pathname = filedialog.askopenfilename(filetypes=[("MS Excel Spreadsheet", ".xlsx .xls .xlsm .xlm")])
    ceo_data = read_excel(ceo_pathname)

    # Open the directory containing reliable SVI data.
    reliable_directory_pathname = filedialog.askdirectory()

    # Open the output MS Excel file.
    output_file = Workbook()
    worksheet = output_file.active
    worksheet.title = sub(r"\s+?", '_', "Split CEO Names")

    # Write the header row.
    worksheet["A1"] = "Number Index"
    worksheet["B1"] = "Full Search Name"
    worksheet["C1"] = "Extracted First Name"
    worksheet["D1"] = "Extracted Last Name"

    for cell in ["A1", "B1", "C1", "D1"]:
        worksheet[cell].font = Font(name="Verdana", size=10, bold=True)
        worksheet[cell].alignment = Alignment(horizontal="left", vertical="center")
        worksheet[cell].number_format = '@'

    next_row = 2  # Keep track of the next empty row.
    set_of_names = set()

    for index, row in ceo_data.iterrows():
        # Check whether the current CEO generated reliable data.
        is_reliable = False
        for filename in listdir(reliable_directory_pathname):
            if filename[-4:] == ".csv" and filename[:3] == "%03d" % (index + 2):
                is_reliable = True
                break

        # If the current CEO generated reliable data, then extract his/her first name and last name.
        if is_reliable:
            full_search_name = row["CEO/Owner/Chairman/President\n(As of 2022-02-26)"]

            # If the name already appears in the set, skip it.
            if full_search_name in set_of_names:
                continue
            else:
                set_of_names.add(full_search_name)

            tokens = [sub(r"[,.()]", '', token) for token in full_search_name.split()]
            extract_result = extract(tokens)

            # Write the data to the output file.
            for cell in ["A%d" % next_row, "B%d" % next_row, "C%d" % next_row, "D%d" % next_row]:
                worksheet[cell].font = Font(name="Verdana", size=10)
                worksheet[cell].alignment = Alignment(horizontal="left", vertical="center")
                worksheet[cell].number_format = '@'
                if not extract_result["reliable"]:
                    worksheet[cell].fill = PatternFill("solid", start_color="ffff99")

            worksheet["A%d" % next_row].alignment = Alignment(horizontal="center", vertical="center")

            worksheet["A%d" % next_row] = "%03d" % (index + 2)
            worksheet["B%d" % next_row] = full_search_name
            worksheet["C%d" % next_row] = extract_result["first_name"]
            worksheet["D%d" % next_row] = extract_result["last_name"]

            next_row += 1

    # Save the output file.
    output_file.save("Split_CEO_Names.xlsx")


if __name__ == "__main__":
    main()
